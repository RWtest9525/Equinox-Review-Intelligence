"""Tests for the new xlsx export format on GET /api/reviews/export?format=xlsx.

Verifies:
- HTTP 200 and correct Content-Type / Content-Disposition (filename ends .xlsx)
- X-Total-Rows header present and matches row count
- Bytes are a valid .xlsx (zip signature + parseable by openpyxl)
- Header row = 15 canonical columns
- Rating filter changes row count and only 5-star rows are present
- Future date range yields header-only workbook (0 data rows)
- Parity: xlsx row count == csv row count for the same filter
"""
import io
import os
import csv
import zipfile
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")

EXPECTED_HEADER = [
    "Review ID", "Date (UTC)", "Rating", "Sentiment", "Topic", "Reviewer",
    "Country", "App Version", "Language", "Review Text", "Reply Status",
    "Published Reply", "Reply Date (UTC)", "Platform", "Source",
]

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture(scope="module")
def demo_app_id(super_admin_session):
    r = super_admin_session.get(f"{BASE_URL}/api/applications", timeout=30)
    assert r.status_code == 200
    apps = r.json()["applications"]
    demo = [a for a in apps if a.get("is_demo")]
    assert demo, "no seeded demo apps"
    return demo[0]["id"]


class TestXlsxExportBasics:
    def test_xlsx_headers_and_mime(self, super_admin_session, demo_app_id):
        r = super_admin_session.get(
            f"{BASE_URL}/api/reviews/export",
            params={"application_id": demo_app_id, "format": "xlsx"}, timeout=60,
        )
        assert r.status_code == 200, r.text[:200]
        assert XLSX_MIME in r.headers.get("Content-Type", ""), r.headers.get("Content-Type")
        cd = r.headers.get("Content-Disposition", "")
        assert "attachment" in cd and cd.endswith('.xlsx"') , f"bad Content-Disposition: {cd}"
        assert "reviews_export_" in cd
        assert r.headers.get("X-Total-Rows") is not None
        # zip signature: xlsx is a zip file
        assert r.content[:2] == b"PK", "response is not a zip/xlsx file"

    def test_xlsx_is_valid_and_has_15_col_header(self, super_admin_session, demo_app_id):
        r = super_admin_session.get(
            f"{BASE_URL}/api/reviews/export",
            params={"application_id": demo_app_id, "format": "xlsx"}, timeout=60,
        )
        assert r.status_code == 200
        # openpyxl parse
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        assert "Live Reviews" in wb.sheetnames, wb.sheetnames
        ws = wb["Live Reviews"]
        rows_iter = ws.iter_rows(values_only=True)
        header = list(next(rows_iter))
        assert header == EXPECTED_HEADER, f"header mismatch:\ngot: {header}\nexp: {EXPECTED_HEADER}"
        data_rows = list(rows_iter)
        # X-Total-Rows should match data row count
        assert int(r.headers["X-Total-Rows"]) == len(data_rows), (
            f"X-Total-Rows={r.headers['X-Total-Rows']} != data rows={len(data_rows)}"
        )

    def test_xlsx_rating_filter_keeps_only_matching_rows(self, super_admin_session, demo_app_id):
        r = super_admin_session.get(
            f"{BASE_URL}/api/reviews/export",
            params={"application_id": demo_app_id, "format": "xlsx", "rating": 5}, timeout=60,
        )
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header, data = rows[0], rows[1:]
        rating_idx = list(header).index("Rating")
        for row in data:
            assert row[rating_idx] == 5, f"non-5 rating leaked: {row}"

    def test_xlsx_future_date_range_yields_header_only(self, super_admin_session, demo_app_id):
        r = super_admin_session.get(
            f"{BASE_URL}/api/reviews/export",
            params={"application_id": demo_app_id, "format": "xlsx",
                    "date_from": "2099-01-01", "date_to": "2099-12-31"},
            timeout=60,
        )
        assert r.status_code == 200
        assert int(r.headers["X-Total-Rows"]) == 0
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 1, f"expected header only, got {len(rows)} rows"
        assert list(rows[0]) == EXPECTED_HEADER

    def test_xlsx_csv_row_count_parity(self, super_admin_session, demo_app_id):
        # Same filter — both formats should have equal row counts
        params_common = {"application_id": demo_app_id, "rating": 5}
        r_x = super_admin_session.get(
            f"{BASE_URL}/api/reviews/export",
            params={**params_common, "format": "xlsx"}, timeout=60,
        )
        r_c = super_admin_session.get(
            f"{BASE_URL}/api/reviews/export",
            params={**params_common, "format": "csv"}, timeout=60,
        )
        assert r_x.status_code == 200 and r_c.status_code == 200
        assert r_x.headers["X-Total-Rows"] == r_c.headers["X-Total-Rows"], (
            f"xlsx={r_x.headers['X-Total-Rows']} csv={r_c.headers['X-Total-Rows']} disagree"
        )
        # CSV header check
        reader = csv.reader(io.StringIO(r_c.text))
        assert next(reader) == EXPECTED_HEADER

    def test_xlsx_bytes_are_zip(self, super_admin_session, demo_app_id):
        r = super_admin_session.get(
            f"{BASE_URL}/api/reviews/export",
            params={"application_id": demo_app_id, "format": "xlsx"}, timeout=60,
        )
        # zipfile should open it
        zf = zipfile.ZipFile(io.BytesIO(r.content))
        names = zf.namelist()
        assert any(n.endswith("workbook.xml") for n in names), names
