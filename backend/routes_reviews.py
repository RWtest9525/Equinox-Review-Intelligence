from fastapi import APIRouter, Depends, Query, Response
from typing import Optional
from deps import db, get_current_user, org_scope
from datetime import datetime, timezone, timedelta
import csv
import io

router = APIRouter(prefix="/api", tags=["reviews"])


def _date_range(date_from: Optional[str], date_to: Optional[str]):
    rng = {}
    if date_from:
        rng["$gte"] = f"{date_from}T00:00:00+00:00"
    if date_to:
        rng["$lte"] = f"{date_to}T23:59:59.999999+00:00"
    return rng or None


@router.get("/reviews")
async def list_reviews(
    user: dict = Depends(get_current_user),
    application_id: Optional[str] = None,
    platform: Optional[str] = None,
    rating: Optional[int] = None,
    sentiment: Optional[str] = None,
    topic: Optional[str] = None,
    country: Optional[str] = None,
    app_version: Optional[str] = None,
    language: Optional[str] = None,
    reply_status: Optional[str] = None,
    quick_filter: Optional[str] = None,
    search: Optional[str] = None,
    days: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort: str = "recent",
    page: int = 1,
    page_size: int = 25,
):
    q = dict(org_scope(user))
    if application_id:
        q["application_id"] = application_id
    if platform and platform in ("google_play", "app_store"):
        q["platform"] = platform
    if rating:
        q["rating"] = rating
    if sentiment:
        q["sentiment"] = sentiment
    if topic:
        q["topic"] = topic
    if country:
        q["country"] = country
    if app_version:
        q["app_version"] = app_version
    if language:
        q["language"] = language
    if reply_status:
        q["reply_status"] = reply_status
    if days:
        q["created_at"] = {"$gte": (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()}
    dr = _date_range(date_from, date_to)
    if dr:
        q["created_at"] = dr

    if quick_filter == "unreplied":
        q["reply_status"] = "unreplied"
    elif quick_filter == "1_star":
        q["rating"] = 1
    elif quick_filter == "2_star":
        q["rating"] = 2
    elif quick_filter == "negative":
        q["sentiment"] = "negative"
    elif quick_filter == "high_priority":
        q["priority"] = "high"
    elif quick_filter == "ai_ready":
        q["reply_status"] = "unreplied"

    if search:
        q["$or"] = [
            {"text": {"$regex": search, "$options": "i"}},
            {"reviewer_name": {"$regex": search, "$options": "i"}},
            {"topic": {"$regex": search, "$options": "i"}},
        ]

    total = await db.reviews.count_documents(q)
    sort_field = ("rating", 1) if sort == "rating_asc" else ("rating", -1) if sort == "rating_desc" else ("created_at", -1)
    cursor = db.reviews.find(q, {"_id": 0}).sort([sort_field]).skip((page - 1) * page_size).limit(page_size)
    reviews = await cursor.to_list(page_size)
    return {"reviews": reviews, "total": total, "page": page, "page_size": page_size}


@router.get("/reviews/export")
async def export_reviews(
    user: dict = Depends(get_current_user),
    application_id: Optional[str] = None,
    platform: Optional[str] = None,
    rating: Optional[int] = None,
    sentiment: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    search: Optional[str] = None,
    format: str = "csv",
):
    q = dict(org_scope(user))
    if application_id:
        q["application_id"] = application_id
    if platform and platform in ("google_play", "app_store"):
        q["platform"] = platform
    if rating:
        q["rating"] = rating
    if sentiment:
        q["sentiment"] = sentiment
    dr = _date_range(date_from, date_to)
    if dr:
        q["created_at"] = dr
    if search:
        q["$or"] = [
            {"text": {"$regex": search, "$options": "i"}},
            {"reviewer_name": {"$regex": search, "$options": "i"}},
        ]
    rows = await db.reviews.find(q, {"_id": 0}).sort([("created_at", -1)]).to_list(20000)

    HEADERS = ["Review ID", "Date (UTC)", "Rating", "Sentiment", "Topic", "Reviewer",
               "Country", "App Version", "Language", "Review Text", "Reply Status",
               "Published Reply", "Reply Date (UTC)", "Platform", "Source"]

    def row_values(r):
        return [
            r.get("external_id") or r.get("id"),
            (r.get("created_at") or "")[:19].replace("T", " "),
            r.get("rating"), r.get("sentiment"), r.get("topic"), r.get("reviewer_name"),
            r.get("country"), r.get("app_version"), (r.get("language") or "").upper(),
            (r.get("text") or "").replace("\n", " ").strip(),
            r.get("reply_status"),
            (r.get("published_reply") or "").replace("\n", " ").strip(),
            (r.get("reply_at") or "")[:19].replace("T", " "),
            r.get("platform"), r.get("source") or ("demo" if r.get("is_demo") else "live"),
        ]

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    truncated = "true" if len(rows) >= 20000 else "false"

    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Live Reviews"
        ws.append(HEADERS)
        header_fill = PatternFill("solid", fgColor="0F172A")
        header_font = Font(bold=True, color="FFFFFF")
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")
        for r in rows:
            ws.append(row_values(r))
        widths = [22, 20, 8, 11, 16, 20, 16, 14, 10, 60, 14, 50, 20, 14, 18]
        for i, wdt in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = wdt
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"
        bio = io.BytesIO()
        wb.save(bio)
        return Response(
            content=bio.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="reviews_export_{stamp}.xlsx"',
                     "X-Total-Rows": str(len(rows)), "X-Truncated": truncated},
        )

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(HEADERS)
    for r in rows:
        w.writerow(row_values(r))
    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="reviews_export_{stamp}.csv"',
                 "X-Total-Rows": str(len(rows)), "X-Truncated": truncated},
    )


@router.get("/reviews/summary")
async def reviews_summary(
    user: dict = Depends(get_current_user),
    application_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    rating: Optional[int] = None,
):
    q = dict(org_scope(user))
    if application_id:
        q["application_id"] = application_id
    if rating:
        q["rating"] = rating
    dr = _date_range(date_from, date_to)
    if dr:
        q["created_at"] = dr
    rows = await db.reviews.find(q, {"_id": 0, "rating": 1, "sentiment": 1}).to_list(20000)
    total = len(rows)
    dist = {i: 0 for i in range(1, 6)}
    sent = {"positive": 0, "neutral": 0, "negative": 0}
    for r in rows:
        dist[r["rating"]] = dist.get(r["rating"], 0) + 1
        sent[r["sentiment"]] = sent.get(r["sentiment"], 0) + 1
    avg = round(sum(r["rating"] for r in rows) / total, 2) if total else 0
    return {
        "total": total, "avg_rating": avg,
        "distribution": [{"stars": s, "count": dist[s], "pct": round(dist[s] / total * 100) if total else 0} for s in range(5, 0, -1)],
        "sentiment": sent,
    }



@router.get("/reviews/filters")
async def review_filter_options(user: dict = Depends(get_current_user)):
    scope = org_scope(user)
    countries = await db.reviews.distinct("country", scope)
    versions = await db.reviews.distinct("app_version", scope)
    topics = await db.reviews.distinct("topic", scope)
    languages = await db.reviews.distinct("language", scope)
    return {
        "countries": sorted(countries),
        "versions": sorted(versions),
        "topics": sorted(topics),
        "languages": sorted(languages),
    }


@router.get("/reviews/{review_id}")
async def get_review(review_id: str, user: dict = Depends(get_current_user)):
    r = await db.reviews.find_one({"id": review_id, **org_scope(user)}, {"_id": 0})
    if not r:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Review not found")
    return {"review": r}


@router.get("/search/global")
async def global_search(q: str = Query(...), user: dict = Depends(get_current_user)):
    scope = org_scope(user)
    rx = {"$regex": q, "$options": "i"}
    reviews = await db.reviews.find({**scope, "text": rx}, {"_id": 0}).limit(5).to_list(5)
    apps = await db.applications.find({**scope, "name": rx}, {"_id": 0}).limit(5).to_list(5)
    comps = await db.competitors.find({**scope, "name": rx}, {"_id": 0}).limit(5).to_list(5)
    return {"reviews": reviews, "applications": apps, "competitors": comps}
